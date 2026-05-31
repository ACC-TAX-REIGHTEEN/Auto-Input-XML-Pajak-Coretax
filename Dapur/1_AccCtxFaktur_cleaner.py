import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def bersihkan_data_faktur_final(input_file, output_file):
    print("--> Memulai pembersihan data")
    
    target_headers = {
        'TGL': 'TGL',                       
        'REFERENSI': 'REFERENSI',            
        'NAMA': 'NAMA',        
        'NPWP': 'NPWP',                
        'No PKP Pelanggan': 'No PKP Pelanggan',    
        'ALAMAT PAJAK 1': 'ALAMAT PAJAK 1',         
        'Negara Pelanggan': 'Negara Pelanggan',       
        'DISC INV': 'DISC INV',              
        'Nilai Faktur': 'Nilai Faktur',             
        'Tax 1 Base Currency': 'Tax 1 Base Currency', 
        'Nama Penjual': 'Nama Penjual'
    }

    urutan_wajib = [
        'TGL', 'REFERENSI', 'NAMA', 'NPWP', 'No PKP Pelanggan', 
        'ALAMAT PAJAK 1', 'Negara Pelanggan', 'DISC INV', 
        'Nilai Faktur', 'Tax 1 Base Currency', 'Nama Penjual', 'IDTKU'
    ]

    try:
        df_raw = pd.read_excel(input_file, header=None, nrows=20, dtype=str)
    except Exception as e:
        print(f"--> Error membaca file: {e}")
        return

    found_mapping = {}      
    max_header_row_idx = 0  
    
    print("--> Sedang mencari posisi kolom...")
    for r_idx, row in df_raw.iterrows():
        for c_idx, cell_value in enumerate(row):
            if pd.isna(cell_value): continue
            clean_val = str(cell_value).strip()
            
            if clean_val in target_headers:
                nama_kolom_baru = target_headers[clean_val]
                if nama_kolom_baru not in found_mapping:
                    found_mapping[nama_kolom_baru] = c_idx
                    if r_idx > max_header_row_idx:
                        max_header_row_idx = r_idx

    if not found_mapping:
        print("--> ERROR: Tidak ada header ditemukan.")
        return

    start_data_row = max_header_row_idx + 1
    df_full = pd.read_excel(input_file, header=None, skiprows=start_data_row, dtype=str)
    df_bersih = pd.DataFrame()
    for nama_baru, index_kolom_asli in found_mapping.items():
        if index_kolom_asli < df_full.shape[1]:
            df_bersih[nama_baru] = df_full.iloc[:, index_kolom_asli]

    if 'NAMA' in df_bersih.columns:
        df_bersih = df_bersih.dropna(subset=['NAMA'])

    print("--> Memformat data...")

    if 'TGL' in df_bersih.columns:
        def paksa_format_indonesia(tgl_str):
            try:
                tgl_str = str(tgl_str).strip()
                if '/' in tgl_str:
                    parts = tgl_str.split('/')
                    if len(parts) == 3:
                        d, m, y = parts
                        if len(y) == 2: 
                            y = "20" + y 
                        return f"{d}/{m}/{y}"
                return tgl_str
            except:
                return tgl_str

        df_bersih['TGL'] = df_bersih['TGL'].apply(paksa_format_indonesia)
        df_bersih['TGL'] = pd.to_datetime(df_bersih['TGL'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')

    cols_angka = ['DISC INV', 'Nilai Faktur', 'Tax 1 Base Currency']
    for col in cols_angka:
        if col in df_bersih.columns:
            df_bersih[col] = df_bersih[col].str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
            df_bersih[col] = pd.to_numeric(df_bersih[col], errors='coerce').fillna(0)

    # C. Buat IDTKU
    if 'NPWP' in df_bersih.columns:
        def buat_idtku(row):
            npwp = str(row['NPWP']).strip()
            if not npwp or npwp.lower() == 'nan': return '000000'
            if npwp == '0000000000000000': return '000000'
            return npwp + '000000'
        df_bersih['IDTKU'] = df_bersih.apply(buat_idtku, axis=1)

    print("--> Mengurutkan kolom sesuai standar...")
    
    kolom_final_tersedia = [kol for kol in urutan_wajib if kol in df_bersih.columns]
    
    df_bersih = df_bersih[kolom_final_tersedia]

    print(f"--> Menyimpan ke {output_file}...")
    df_bersih.to_excel(output_file, index=False, sheet_name='Data Bersih')
    
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[col_letter].width = (max_length + 2)
        wb.save(output_file)
    except: pass

    print("--> Selesai! Kolom sudah urut kembali.")

if __name__ == "__main__":
    nama_input = 'AccCtxFaktur.xls'
    nama_output = 'AccCtxFaktur_temp.xlsx'
    bersihkan_data_faktur_final(nama_input, nama_output)        
