import openpyxl
from datetime import datetime

def perbaiki_format_tanggal(nama_file, nama_sheet, nama_kolom):
    print(f"--> Sedang memproses file: {nama_file}...")
    
    try:
        wb = openpyxl.load_workbook(nama_file)
        
        if nama_sheet not in wb.sheetnames:
            print(f"--> Error: Sheet '{nama_sheet}' tidak ditemukan.")
            return

        ws = wb[nama_sheet]

        kolom_target_index = None
        for cell in ws[1]: 
            if cell.value == nama_kolom:
                kolom_target_index = cell.column
                break
        
        if kolom_target_index is None:
            print(f"--> Error: Kolom '{nama_kolom}' tidak ditemukan di baris pertama.")
            return

        print(f"--> Kolom '{nama_kolom}' ditemukan di indeks ke-{kolom_target_index}.")

        jumlah_perbaikan = 0
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=kolom_target_index)
            nilai_asli = cell.value
            if isinstance(nilai_asli, str):
                try:
                    dt_object = datetime.strptime(nilai_asli, '%d/%m/%Y')
                    
                    cell.value = dt_object
                    
                    cell.number_format = 'DD/MM/YYYY'
                    
                    jumlah_perbaikan += 1
                except ValueError:
                    pass

        wb.save(nama_file)
        print(f"--> Selesai! {jumlah_perbaikan} data tanggal berhasil diperbaiki.")
        print(f"--> File disimpan kembali ke: {nama_file}")

    except Exception as e:
        print(f"--> Terjadi kesalahan: {e}")

if __name__ == "__main__":
    file_excel = "MkNwFile_temp.xlsx"
    sheet_target = "Faktur"
    kolom_target = "Tanggal Faktur"

    perbaiki_format_tanggal(file_excel, sheet_target, kolom_target)
