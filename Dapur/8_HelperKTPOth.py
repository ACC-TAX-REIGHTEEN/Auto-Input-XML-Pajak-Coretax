import openpyxl

file_excel = 'MkNwFile_temp.xlsx'
file_output = 'MkNwFile_temp.xlsx'
file_ktp = 'KTP.txt'
file_ktp_oth = 'KTP-OTH.txt'

with open(file_ktp, 'r') as f:
    ktp_data = set(line.strip() for line in f if line.strip())

with open(file_ktp_oth, 'r') as f:
    ktp_oth_data = set(line.strip() for line in f if line.strip())

wb = openpyxl.load_workbook(file_excel)

if 'Faktur' in wb.sheetnames:
    ws = wb['Faktur']
    
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    col_npwp = headers.get('NPWP/NIK Pembeli')
    col_dokumen = headers.get('Nomor Dokumen Pembeli')
    col_id_tku = headers.get('ID TKU Pembeli')
    col_jenis = headers.get('Jenis ID Pembeli')

    if col_npwp and col_dokumen and col_id_tku and col_jenis:
        for row in ws.iter_rows(min_row=2):
            cell_npwp = row[col_npwp - 1]
            val_npwp = str(cell_npwp.value).strip() if cell_npwp.value else ""

            if val_npwp in ktp_data:
                row[col_dokumen - 1].value = val_npwp
                cell_npwp.value = "0000000000000000"
                row[col_id_tku - 1].value = "000000"
                row[col_jenis - 1].value = "National ID"
            
            elif val_npwp in ktp_oth_data:
                cell_npwp.value = "0000000000000000"
                row[col_jenis - 1].value = "Other ID"
                row[col_dokumen - 1].value = "0000000000000000"
                row[col_id_tku - 1].value = "000000"

wb.save(file_output)
print(f"--> Selesai. File tersimpan di {file_output}")