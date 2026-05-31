import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

file_efaktur = 'AccEFaktur_temp.xlsx'     
file_ctx = 'AccCtxFaktur_temp.xlsx'     
file_target = 'MkNwFile_temp.xlsx'       

try:
    df_efaktur = pd.read_excel(file_efaktur)
    df_ctx = pd.read_excel(file_ctx)
except Exception as e:
    print(f"--> Error membaca file: {e}")
    exit()

data_detail = df_efaktur['No.Inv'].astype(str).str.strip().tolist()
data_faktur = df_ctx['REFERENSI'].astype(str).str.strip().tolist()

max_row = max(len(data_detail), len(data_faktur))

faktur_map = {}
for idx, val in enumerate(data_faktur):
    if val not in faktur_map:
        faktur_map[val] = idx + 1 

data_hasil = []
for item in data_detail:
    urutan_ketemu = faktur_map.get(item, 0)
    data_hasil.append(urutan_ketemu)

try:
    wb = openpyxl.load_workbook(file_target)
    
    if 'Kalkulasi' in wb.sheetnames:
        ws = wb['Kalkulasi']
        ws.delete_rows(1, ws.max_row + 1)
    else:
        ws = wb.create_sheet('Kalkulasi')

    headers = ["DETAIL", "FAKTUR", "URUTAN", "HASIL"]
    ws.append(headers)

    for i in range(max_row):
        val_a = data_detail[i] if i < len(data_detail) else None
        val_b = data_faktur[i] if i < len(data_faktur) else None
        val_c = i + 1 
        val_d = data_hasil[i] if i < len(data_hasil) else None

        ws.cell(row=i+2, column=1, value=val_a)
        ws.cell(row=i+2, column=2, value=val_b)
        ws.cell(row=i+2, column=3, value=val_c)
        ws.cell(row=i+2, column=4, value=val_d)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 3

    wb.save(file_target)
    print("--> Selesai. Logika XLOOKUP diterapkan.")

except Exception as e:
    print(f"--> Error menyimpan file: {e}")
