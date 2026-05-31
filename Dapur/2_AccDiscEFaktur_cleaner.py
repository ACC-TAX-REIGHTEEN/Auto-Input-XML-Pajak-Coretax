import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def proses_data_faktur_dynamic(input_file, output_file):
    print(f"-->  Mulai memproses file: {input_file} ---")

    target_headers = {
        'No.Inv': 'No.Inv',
        'Tgl Faktur': 'Tgl Faktur',
        'Nama Pelanggan': 'Nama Pelanggan',
        'Alamat Pajak 1 Pelanggan': 'Alamat Pajak 1 Pelanggan',
        'Kota': 'Kota',
        'No. Barang': 'No. Barang',
        'Nama Barang': 'Nama Barang',
        'Qty': 'Qty',
        'No. Barang': 'No. Barang',
        'H.Jual': 'H.Jual',
        'Harga Sat': 'Harga Sat',
        'Total Harga': 'Total Harga',
        'Nilai Faktur': 'Nilai Faktur',
        'Tax 1': 'Tax 1',
        'Discount Faktur': 'Discount Faktur',
        'No. Pelanggan': 'No. Pelanggan',
        'Nomor Pajak Pelanggan': 'Nomor Pajak Pelanggan'
    }

    urutan_output = [
        'No.Inv', 'Tgl Faktur', 'Nama Pelanggan', 'Alamat Pajak 1 Pelanggan',
        'Kota', 'No. Barang', 'Nama Barang', 'Qty', 'No. Barang',
        'H.Jual', 'Harga Sat', 'Total Harga', 'Nilai Faktur',
        'Tax 1', 'Discount Faktur', 'No. Pelanggan', 'Nomor Pajak Pelanggan',
        'HJTNP',         
        'DISC. TANPA',    
        'TOTAL QTY',      
        'DISC. SATUAN',    
        'DISC. ITEM'       
    ]

    print("--> Sedang memindai posisi kolom...")
    try:
        df_raw = pd.read_excel(input_file, header=None, nrows=20, dtype=str)
    except Exception:
        try:
            df_raw = pd.read_csv(input_file, header=None, nrows=20, dtype=str)
        except Exception as e:
            print(f"--> Error fatal: Tidak bisa membaca file. {e}")
            return

    found_mapping = {}      
    max_header_row_idx = 0  

    for r_idx, row in df_raw.iterrows():
        for c_idx, cell_value in enumerate(row):
            if pd.isna(cell_value): continue
            clean_val = str(cell_value).strip()
            
            if clean_val in target_headers:
                nama_kolom_baru = target_headers[clean_val]
                if nama_kolom_baru not in found_mapping:
                    found_mapping[nama_kolom_baru] = c_idx
                    if r_idx > max_header_row_idx:
                        max_header_row_idx = r_idx

    if not found_mapping:
        print("--> ERROR: Tidak ada header yang dikenali. Cek ejaan di 'target_headers'.")
        return
    else:
        print(f"--> Header ditemukan. Data dimulai setelah baris {max_header_row_idx + 1}.")

    try:
        df_full = pd.read_excel(input_file, header=None, skiprows=max_header_row_idx + 1, dtype=str)
    except:
        df_full = pd.read_csv(input_file, header=None, skiprows=max_header_row_idx + 1, dtype=str)

    df_clean = pd.DataFrame()
    for nama_baru, index_kolom_asli in found_mapping.items():
        if index_kolom_asli < df_full.shape[1]:
            df_clean[nama_baru] = df_full.iloc[:, index_kolom_asli]

    if 'No.Inv' in df_clean.columns:
        df_clean = df_clean.dropna(subset=['No.Inv'])

    print("--> Membersihkan format data...")

    cols_suffix = ['No.Inv', 'No. Barang', 'Qty', 'Discount Faktur', 'No. Pelanggan', 'Nomor Pajak Pelanggan']
    for col in cols_suffix:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).str.replace(r'\.0$', '', regex=True)
            df_clean[col] = df_clean[col].str.replace(r',00$', '', regex=True)
            df_clean[col] = df_clean[col].str.strip()

    if 'Tgl Faktur' in df_clean.columns:
        def perbaiki_tanggal(tgl):
            try:
                if pd.isna(tgl): return tgl
                parts = str(tgl).split('/')
                if len(parts) == 3:
                    d, m, y = parts
                    if len(y) == 2: y = "20" + y
                    return f"{d}/{m}/{y}"
                return tgl
            except: return tgl
        df_clean['Tgl Faktur'] = df_clean['Tgl Faktur'].apply(perbaiki_tanggal)

    cols_angka = ['H.Jual', 'Harga Sat', 'Total Harga', 'Nilai Faktur']
    cols_angka_tambahan = ['Discount Faktur', 'Qty']
    for col in cols_angka + cols_angka_tambahan:
        if col in df_clean.columns:
            temp_col = df_clean[col].astype(str)
            temp_col = temp_col.str.replace(',', '.', regex=False)
            df_clean[col] = pd.to_numeric(temp_col, errors='coerce').fillna(0)

    print("--> Melakukan kalkulasi tambahan...")

    if 'H.Jual' in df_clean.columns:
        df_clean['HJTNP'] = df_clean['H.Jual'] / 1.11
        cols_angka.append('HJTNP')

    if 'Discount Faktur' in df_clean.columns and 'Qty' in df_clean.columns and 'No.Inv' in df_clean.columns:
        
        df_clean['DISC. TANPA'] = df_clean['Discount Faktur'] / 1.11
        df_clean['TOTAL QTY'] = df_clean.groupby('No.Inv')['Qty'].transform('sum')
        df_clean['DISC. SATUAN'] = df_clean['DISC. TANPA'] / df_clean['TOTAL QTY']
        df_clean['DISC. SATUAN'] = df_clean['DISC. SATUAN'].fillna(0)
        df_clean['DISC. ITEM'] = df_clean['DISC. SATUAN'] * df_clean['Qty']

        cols_angka.extend(['DISC. TANPA', 'TOTAL QTY', 'DISC. SATUAN', 'DISC. ITEM'])

    print("--> Mengurutkan kolom...")
    kolom_final = [k for k in urutan_output if k in df_clean.columns]
    df_clean = df_clean[kolom_final]

    print(f"--> Menyimpan ke {output_file}...")
    df_clean.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for i, column in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        header_cell = column[0]
        header_text = str(header_cell.value) if header_cell.value else ""
        max_length = len(header_text)
        
        for cell in column[1:]: 
            try:
                if cell.value is not None:
                    if header_text in cols_angka:
                        cell.number_format = '#,##0.00'
                    
                    val_len = len(str(cell.value))
                    if val_len > max_length: max_length = val_len
            except: pass
        
        adj_width = max_length + 2
        if adj_width > 50: adj_width = 50
        ws.column_dimensions[col_letter].width = adj_width

    wb.save(output_file)
    print(f"-->  SELESAI! Hasil tersimpan di: {output_file} ---")

if __name__ == "__main__":
    file_masuk = 'EFaktur.xls'
    file_keluar = 'AccEFaktur_temp.xlsx'
    
    if os.path.exists(file_masuk):
        proses_data_faktur_dynamic(file_masuk, file_keluar)
    else:
        print(f"File '{file_masuk}' tidak ditemukan. Pastikan nama file benar.")
