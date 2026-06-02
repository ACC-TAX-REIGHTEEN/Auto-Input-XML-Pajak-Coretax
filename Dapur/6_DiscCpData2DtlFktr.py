import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def load_helper_list(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    return []

def main():
    print("--> Sedang memuat file Excel...")
    try:
        df_kalkulasi = pd.read_excel('MkNwFile_temp.xlsx', sheet_name='Kalkulasi')
        df_acc = pd.read_excel('AccEFaktur_temp.xlsx') 
    except Exception as e:
        print(f"--> Error membaca file Excel: {e}")
        return

    print("--> Sedang memuat file Helper...")
    helpers_kode = {
    	'160100': load_helper_list('Helper_160100.txt'),
    	'200104': load_helper_list('Helper_200104.txt'),
        '340300': load_helper_list('Helper_340300.txt'),
        '401100': load_helper_list('Helper_401100.txt'),
        '401300': load_helper_list('Helper_401300.txt'),
        '400800': load_helper_list('Helper_400800.txt'),
        '830900': load_helper_list('Helper_830900.txt'),
        '842100': load_helper_list('Helper_842100.txt'),
    }
    
    helper_a = load_helper_list('Helper_A.txt')
    helper_b = load_helper_list('Helper_B.txt')
    helper_del = load_helper_list('Helper_Del.txt')

    jumlah_baris = len(df_acc)
    df_result = pd.DataFrame(index=range(jumlah_baris))

    print("--> Sedang memproses data...")

    df_result['Nama Barang/Jasa'] = df_acc['Nama Barang']

    def get_kode_barang(nama_barang):
        for kode, daftar_barang in helpers_kode.items():
            if nama_barang in daftar_barang:
                return kode
        return ""
    
    df_result['Kode Barang Jasa'] = df_result['Nama Barang/Jasa'].apply(get_kode_barang)
)
    def get_barang_jasa_group(kode_barang):
        kode_str = str(kode_barang)
        if kode_str in helper_a:
            return "A"
        elif kode_str in helper_b:
            return "B"
        return ""
    
    df_result['Barang/Jasa'] = df_result['Kode Barang Jasa'].apply(get_barang_jasa_group)

    def get_satuan(group):
        if group == "A":
            return "UM.0021"
        elif group == "B":
            return "UM.0030"
        return ""
    
    df_result['Nama Satuan Ukur'] = df_result['Barang/Jasa'].apply(get_satuan)
    df_result['Baris'] = df_kalkulasi['HASIL']
    df_result['Harga Satuan'] = df_acc['HJTNP']
    df_result['Jumlah Barang Jasa'] = df_acc['Qty']
    
    try:
        df_result['Total Diskon'] = df_acc['DISC. ITEM']
    except KeyError:
        print("--> PERINGATAN: Kolom 'DISC. ITEM' tidak ditemukan. Menggunakan 0 sebagai default.")
        df_result['Total Diskon'] = 0

    df_result['DPP'] = (df_result['Harga Satuan'] * df_result['Jumlah Barang Jasa']) - df_result['Total Diskon']
    df_result['DPP Nilai Lain'] = (11/12) * df_result['DPP']
    df_result['Tarif PPN'] = 12
    df_result['PPN'] = (df_result['DPP Nilai Lain'] * df_result['Tarif PPN']) / 100
    df_result['Tarif PPnBM'] = 0
    df_result['PPnBM'] = (df_result['Tarif PPnBM'] * df_result['DPP Nilai Lain']) / 100

    kolom_final = [
        'Baris', 'Barang/Jasa', 'Kode Barang Jasa', 'Nama Barang/Jasa', 
        'Nama Satuan Ukur', 'Harga Satuan', 'Jumlah Barang Jasa', 'Total Diskon',
        'DPP', 'DPP Nilai Lain', 'Tarif PPN', 'PPN', 'Tarif PPnBM', 'PPnBM'
    ]
    
    df_final = df_result[kolom_final]

    if helper_del:
        print(f"--> Mengecek Helper_Del.txt ({len(helper_del)} item)...")
        jumlah_awal = len(df_final)
        
        df_final = df_final[~df_final['Nama Barang/Jasa'].isin(helper_del)]
        
        jumlah_akhir = len(df_final)
        print(f"--> Berhasil menghapus {jumlah_awal - jumlah_akhir} baris berdasarkan nama.")
    else:
        print("--> Helper_Del.txt kosong atau tidak ditemukan, tidak ada data dihapus.")

    print("--> Mengecek harga minus di kolom 'Harga Satuan'...")
    jumlah_sebelum_minus = len(df_final)
    
    df_final = df_final[df_final['Harga Satuan'] >= 0]
    
    jumlah_sesudah_minus = len(df_final)
    selisih_minus = jumlah_sebelum_minus - jumlah_sesudah_minus
    
    if selisih_minus > 0:
        print(f"--> Berhasil menghapus {selisih_minus} baris yang memiliki harga minus.")
    else:
        print("--> Aman! Tidak ditemukan harga minus.")

    print("--> Menyimpan ke MkNwFile_temp.xlsx...")
    output_file = 'MkNwFile_temp.xlsx'
    
    try:
        wb = load_workbook(output_file)
        if 'DetailFaktur' in wb.sheetnames:
            del wb['DetailFaktur']
            wb.save(output_file)
        wb.close()
    except Exception as e:
        print(f"--> Info: File baru atau sheet belum ada ({e})")

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df_final.to_excel(writer, sheet_name='DetailFaktur', index=False)
        
        print("--> Menerapkan auto-fit kolom...")
        try:
            worksheet = writer.sheets['DetailFaktur']
            for column in worksheet.columns:
                max_length = 0
                col_idx = column[0].column
                
                if isinstance(col_idx, int):
                    column_letter = get_column_letter(col_idx)
                else:
                    column_letter = col_idx 
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"--> Warning: Auto-fit gagal, tapi data tersimpan. Error: {e}")

    print("--> Selesai! File MkNwFile_temp.xlsx sheet DetailFaktur telah diperbarui.")

if __name__ == "__main__":
    main()